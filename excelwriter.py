from openpyxl import Workbook, load_workbook
from os.path import exists


class ExcelWriter:

    def __init__(self, errorHandler):
        self.filePath = None
        self.errorHandler = errorHandler

    def setFilePath(self, filePath):
        self.filePath = filePath

    def saveRecord(self, record):
        try:
            if not exists(self.filePath):
                self.createWorkbook()
            wb = load_workbook(self.filePath)
            sheet = wb.active
            row = self.createRow(record)
            sheet.append(row)
            wb.save(self.filePath)
        except PermissionError:
            self.errorHandler(
                """Error: Submit Failed. Either Excel file is open in another application,\nor you do not have permission to write to this file."""
            )

    def createWorkbook(self):
        wb = Workbook()
        sheet = wb.active
        sheet.title = "BBB_Data"
        sheet.append(self.getColumnHeaders(wb))
        wb.save(self.filePath)

    def getColumnHeaders(self, wb):
        headers = [
            "Rat #",
            "Week",
            "Date",
            "Tester",
            "Score Left",
            "Score Right",

            "left Hip",
            "left Knee",
            "left Ankle",
            "right Hip",
            "right Knee",
            "right Ankle",
            "trunk Side",
            "trunk Prop",
            "abdomen",
            "left Sweep",
            "right Sweep",
            "left Support",
            "right Support",
            "stepping DorsalLeft",
            "stepping PlantarLeft",
            "stepping DorsalRight",
            "stepping PlantarRight",
            "coordination",
            "left Toe",
            "right Toe",
            "initial Contact Left",
            "initial Contact Right",
            "liftOff Left",
            "liftOff Right",
            "trunk Instability",
            "tail"
        ]
        return headers

    def createRow(self, record):
        row = [
            record.ratNumber,
            record.week,
            record.date,
            record.tester,
            record.scoreLeft,
            record.scoreRight,

            record.leftHip.name,
            record.leftKnee.name,
            record.leftAnkle.name,
            record.rightHip.name,
            record.rightKnee.name,
            record.rightAnkle.name,
            record.trunkSide.name,
            record.trunkProp.name,
            record.abdomen.name,
            record.leftSweep.name,
            record.rightSweep.name,
            record.leftSupport.name,
            record.rightSupport.name,
            record.steppingDorsalLeft.name,
            record.steppingPlantarLeft.name,
            record.steppingDorsalRight.name,
            record.steppingPlantarRight.name,
            record.coordination.name,
            record.leftToe.name,
            record.rightToe.name,
            record.initialContactLeft.name,
            record.initialContactRight.name,
            record.liftOffLeft.name,
            record.liftOffRight.name,
            record.trunkInstability.name,
            record.tail.name
        ]
        return row
